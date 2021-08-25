import pandas as pd
import numpy as np
from nltk.metrics.distance import jaccard_distance
from nltk.util import ngrams
from openpyxl import load_workbook

import config_asta as cfa
import db_functions as dbf


def insert_players_first_time():

    players = pd.read_excel(io=cfa.QUOTAZIONI,
                            sheet_name="Tutti",
                            usecols=[1, 2, 3, 4],
                            engine='openpyxl')

    for x in range(1, len(players)):
        roles, name, team, price = players.iloc[x].values
        dbf.db_insert(table='players',
                      columns=['name', 'team', 'roles', 'price', 'status'],
                      values=[name.strip().title(), team.strip(),
                              roles.strip(), int(price), 'FREE'],
                      database=cfa.DB_NAME)


def jaccard_result(name_to_fix: str, all_options: list, ngrams_lenght: int):

    name_to_correct = name_to_fix.lower().replace(' ', '')
    n_in = set(ngrams(name_to_correct, ngrams_lenght))

    out_opts = [pl.lower().replace(' ', '') for pl in all_options]
    n_outs = [set(ngrams(pl, ngrams_lenght)) for pl in out_opts]

    distances = [jaccard_distance(n_in, n_out) for n_out in n_outs]

    if len(set(distances)) == 1 and distances[0] == 1:
        return jaccard_result(name_to_correct, all_options, ngrams_lenght-1)
    else:
        return np.array(all_options)[np.argsort(distances)][:3]


def offer_is_too_low(player_name: str, offer_value: int):
    base_pr = dbf.db_select(table='players',
                            columns=['price'],
                            where=f'name = "{player_name}"',
                            database=cfa.DB_NAME)[0]
    return (True, base_pr) if offer_value < base_pr else (False, base_pr)


def update_excel(winner_team, pl_name, pl_team, pl_roles, pl_price):

    wb = load_workbook(cfa.FILEPATH)
    ws = wb.active

    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df.drop(0, inplace=True)

    col = list(df.columns).index(winner_team) + 1
    row = df[winner_team][df[winner_team].isna()].index[2] + 1

    ws.cell(row=row, column=col).value = pl_name
    ws.cell(row=row, column=col + 1).value = pl_team
    ws.cell(row=row, column=col + 2).value = pl_roles
    ws.cell(row=row, column=col + 3).value = pl_price

    # data = zip([cfa.fill1, cfa.fill2, cfa.fill3, cfa.fill4,
    #             cfa.fill5, cfa.fill6, cfa.fill7, cfa.fill8],
    #            range(1, 33, 4))
    #
    # for fill, i in data:
    #     for col in ws.iter_cols(min_col=i, max_col=i + 3,
    #                             min_row=None, max_row=None):
    #         for cell in col:
    #             cell.fill = fill

    wb.save(cfa.FILEPATH)


def start_asta():
    print('Type PLAYER_NAME, FANTATEAM, PRICE:')

    try:
        nm, tm, pr = input().split(',')
        pr = int(pr.strip())
    except ValueError:
        print('WRONG FORMAT')
        return start_asta()

    all_nm = dbf.db_select(table='players',
                           columns=['name'],
                           where='status = "FREE"',
                           database=cfa.DB_NAME)
    names = jaccard_result(name_to_fix=nm, all_options=all_nm, ngrams_lenght=3)

    all_tm = dbf.db_select(table='teams',
                           columns=['team_name'],
                           where='',
                           database=cfa.DB_NAME)
    tm = jaccard_result(name_to_fix=tm, all_options=all_tm, ngrams_lenght=3)[0]

    for opt in names:
        real_team, roles = dbf.db_select(table='players',
                                         columns=['team', 'roles'],
                                         where=f'name = "{opt}"',
                                         database=cfa.DB_NAME)[0]
        real_team = real_team[:3].upper()

        print(f'{opt} ({real_team}\t{roles})\t\t{tm}\t\t{pr}')
        print(('ENTER to confirm, SPACE to cancel, '
               'any letter if player is wrong.'))
        answer = input()
        if not answer:
            too_low, min_price = offer_is_too_low(player_name=opt,
                                                  offer_value=pr)
            if too_low:
                print(f'DENIED. Min price is {min_price}.')
            else:
                dbf.db_update(table='players',
                              columns=['status'],
                              values=[tm],
                              where=f'name = "{opt}"',
                              database=cfa.DB_NAME)
                update_excel(winner_team=tm, pl_name=opt, pl_team=real_team,
                             pl_roles=roles, pl_price=pr)
                with open('logs_asta.txt', 'a') as f:
                    f.write(f'{opt}, {real_team}, {roles}, {tm}, {pr}\n')
            return start_asta()
        elif answer == ' ':
            return start_asta()
        else:
            continue


# insert_players_first_time()
start_asta()
